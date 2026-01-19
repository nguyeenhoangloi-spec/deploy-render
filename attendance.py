from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, Response
from connect_postgres import get_connection
from auth import login_required, role_required
import os
import qrcode
from datetime import datetime
from settings import get_all_settings
import socket
import base64, json as _json
import numpy as np
import cv2, pickle
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

attendance_bp = Blueprint("attendance", __name__)

QR_DIR = os.path.join(os.path.dirname(__file__), "outputs", "qrcodes")
os.makedirs(QR_DIR, exist_ok=True)

from flask import send_from_directory

@attendance_bp.route("/qrcode/<path:filename>")
def qrcode_image(filename):
    return send_from_directory(QR_DIR, filename)

@attendance_bp.route("/")
@login_required
def index():
    conn = get_connection(); cur = conn.cursor()
    cur.execute("""
        SELECT s.id, c.code, sub.code, s.session_date, s.time_slot, s.room, s.status
        FROM attendance_sessions s
        JOIN classes c ON s.class_id=c.id
        JOIN subjects sub ON s.subject_id=sub.id
        ORDER BY s.session_date DESC, s.id DESC
    """)
    rows = cur.fetchall(); cur.close(); conn.close()
    return render_template("attendance_list.html", rows=rows)

@attendance_bp.route("/new", methods=["GET", "POST"])
@role_required("admin", "lecturer")
def create():
    conn = get_connection(); cur = conn.cursor()
    # Lọc danh sách lớp theo quyền giảng viên
    role = session.get("role", "lecturer")
    user_id = session.get("user_id")
    if role == "lecturer" and user_id:
        cur.execute(
            """
            SELECT c.id, c.code, c.name
            FROM classes c JOIN instructors_classes ic ON ic.class_id=c.id
            WHERE ic.user_id=%s ORDER BY c.code
            """,
            (user_id,)
        )
    else:
        cur.execute("SELECT id, code, name FROM classes ORDER BY code")
    classes = cur.fetchall()
    cur.execute("SELECT id, code, name FROM subjects ORDER BY code"); subjects = cur.fetchall()
    if request.method == "POST":
        class_id = request.form.get("class_id")
        subject_id = request.form.get("subject_id")
        session_date = request.form.get("session_date")
        time_slot = request.form.get("time_slot")
        room = request.form.get("room")
        # Xác thực class_id thuộc quyền giảng viên (nếu là giảng viên)
        try:
            if role == "lecturer" and user_id and class_id:
                cur2 = conn.cursor()
                cur2.execute("SELECT 1 FROM instructors_classes WHERE class_id=%s AND user_id=%s", (class_id, user_id))
                if not cur2.fetchone():
                    cur2.close()
                    flash("Bạn không được phân công quản lý lớp đã chọn", "error")
                    return render_template("attendance_form.html", classes=classes, subjects=subjects)
                cur2.close()
        except Exception:
            pass
        # generate short session code
        import secrets
        session_code = secrets.token_hex(4).upper()
        try:
            cur.execute("""
                INSERT INTO attendance_sessions (class_id, subject_id, session_date, time_slot, room, status, session_code, created_by)
                VALUES (%s, %s, %s, %s, %s, 'created', %s, %s)
            """, (class_id, subject_id, session_date, time_slot, room, session_code, None))
            conn.commit()
            # Generate QR using public base URL if configured, else try auto-detect LAN IP
            settings = get_all_settings()
            public_base = (settings.get("public_base_url") or "").strip()
            if public_base:
                if public_base.endswith("/"):
                    public_base = public_base[:-1]
                join_url = f"{public_base}{url_for('attendance.mobile_join', session_code=session_code)}"
            else:
                # Auto-detect primary LAN IP (non-loopback)
                try:
                    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                    s.connect(("8.8.8.8", 80))
                    local_ip = s.getsockname()[0]
                    s.close()
                    join_url = f"http://{local_ip}:5000" + url_for("attendance.mobile_join", session_code=session_code)
                except Exception:
                    # Fallback to Flask external URL (may be 127.0.0.1)
                    join_url = url_for("attendance.mobile_join", session_code=session_code, _external=True)
            img = qrcode.make(join_url)
            qr_path = os.path.join(QR_DIR, f"{session_code}.png")
            # Lưu ảnh QR qua stream để phù hợp type checker
            with open(qr_path, "wb") as fp:
                img.save(fp)
            cur = conn.cursor()
            cur.execute("UPDATE attendance_sessions SET qr_code_path=%s WHERE session_code=%s", (qr_path, session_code))
            conn.commit(); cur.close()
            flash("Đã tạo buổi học", "success")
            return redirect(url_for("attendance.index"))
        except Exception as e:
            conn.rollback(); flash(f"Lỗi tạo buổi học: {e}", "error")
    cur.close(); conn.close()
    return render_template("attendance_form.html", classes=classes, subjects=subjects)

@attendance_bp.route("/<int:session_id>")
@login_required
def detail(session_id):
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT id, class_id, subject_id, session_date, time_slot, room, status, session_code, qr_code_path, start_time, end_time FROM attendance_sessions WHERE id=%s", (session_id,))
    sess = cur.fetchone()
    if not sess:
        cur.close(); conn.close(); return redirect(url_for("attendance.index"))
    class_id = sess[1]; subject_id = sess[2]
    # Lấy mã lớp và mã môn để hiển thị
    try:
        cur.execute("SELECT code FROM classes WHERE id=%s", (class_id,))
        row_cls = cur.fetchone(); class_code = row_cls[0] if row_cls else None
    except Exception:
        class_code = None
    try:
        cur.execute("SELECT code FROM subjects WHERE id=%s", (subject_id,))
        row_sub = cur.fetchone(); subject_code = row_sub[0] if row_sub else None
    except Exception:
        subject_code = None
    # Nếu có đăng ký môn học theo sinh viên, ưu tiên danh sách đó; nếu không có, lấy theo lớp
    has_subject_students = False
    try:
        cur.execute("SELECT EXISTS(SELECT 1 FROM student_subjects WHERE subject_id=%s)", (subject_id,))
        row = cur.fetchone()
        has_subject_students = bool(row[0]) if row is not None else False
    except Exception:
        # Nếu bảng chưa tồn tại hoặc lỗi, fallback về danh sách theo lớp
        has_subject_students = False
    if has_subject_students:
        try:
            cur.execute(
                """
                SELECT s.id, s.student_code, s.fullname
                FROM student_subjects ss
                JOIN students s ON s.id=ss.student_id
                WHERE ss.subject_id=%s AND s.class_id=%s
                ORDER BY s.fullname
                """,
                (subject_id, class_id)
            )
            students = cur.fetchall()
        except Exception:
            # Fallback khi truy vấn lỗi
            cur.execute("SELECT s.id, s.student_code, s.fullname FROM students s WHERE s.class_id=%s ORDER BY s.fullname", (class_id,))
            students = cur.fetchall()
    else:
        cur.execute("SELECT s.id, s.student_code, s.fullname FROM students s WHERE s.class_id=%s ORDER BY s.fullname", (class_id,))
        students = cur.fetchall()
    # attendance records
    cur.execute("SELECT student_id, status, confidence, device, captured_at FROM attendance_records WHERE session_id=%s", (session_id,))
    records = cur.fetchall()
    cur.close(); conn.close()
    rec_map = {r[0]: r for r in records}
    # Phân tách danh sách đã/chưa điểm danh để hiển thị riêng
    marked_students = []  # (id, code, fullname, status, confidence, device, captured_at)
    pending_students = [] # (id, code, fullname)
    for s in students:
        sid, scode, sname = s[0], s[1], s[2]
        if sid in rec_map:
            r = rec_map[sid]
            marked_students.append((sid, scode, sname, r[1], r[2], r[3], r[4]))
        else:
            pending_students.append((sid, scode, sname))
    return render_template(
        "attendance_detail.html",
        session=sess,
        class_code=class_code,
        subject_code=subject_code,
        students=students,
        records=rec_map,
        marked_students=marked_students,
        pending_students=pending_students,
    )

@attendance_bp.route("/<int:session_id>/regen_qr", methods=["POST"])
@role_required("admin", "lecturer")
def regen_qr(session_id):
    # Lấy session_code
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT session_code FROM attendance_sessions WHERE id=%s", (session_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close();
        flash("Không tìm thấy phiên", "error")
        return redirect(url_for("attendance.detail", session_id=session_id))
    session_code = row[0]
    # Tạo URL theo public_base_url nếu có
    settings = get_all_settings()
    public_base = (settings.get("public_base_url") or "").strip()
    if public_base:
        if public_base.endswith("/"):
            public_base = public_base[:-1]
        join_url = f"{public_base}{url_for('attendance.mobile_join', session_code=session_code)}"
    else:
        # Auto-detect LAN IP when base URL not configured
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            local_ip = s.getsockname()[0]
            s.close()
            join_url = f"http://{local_ip}:5000" + url_for("attendance.mobile_join", session_code=session_code)
        except Exception:
            join_url = url_for("attendance.mobile_join", session_code=session_code, _external=True)
    # Sinh lại QR
    import qrcode, os
    img = qrcode.make(join_url)
    qr_path = os.path.join(QR_DIR, f"{session_code}.png")
    with open(qr_path, "wb") as fp:
        img.save(fp)
    # Cập nhật đường dẫn QR trong DB
    try:
        cur.execute("UPDATE attendance_sessions SET qr_code_path=%s WHERE id=%s", (qr_path, session_id))
        conn.commit();
        flash("Đã tạo lại QR theo Public Base URL", "success")
    except Exception:
        conn.rollback(); flash("Lỗi tạo lại QR", "error")
    finally:
        cur.close(); conn.close()
    return redirect(url_for("attendance.detail", session_id=session_id))

# Hủy (xóa) điểm danh khi bấm lộn để trả sinh viên về trạng thái "chưa điểm danh"
@attendance_bp.route("/<int:session_id>/unmark", methods=["POST"])
@role_required("admin", "lecturer")
def unmark(session_id):
    student_id = request.form.get("student_id")
    if not student_id:
        return redirect(url_for("attendance.detail", session_id=session_id))
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("DELETE FROM attendance_records WHERE session_id=%s AND student_id=%s", (session_id, student_id))
        conn.commit();
        flash("Đã hủy điểm danh cho sinh viên", "success")
    except Exception:
        conn.rollback();
        flash("Lỗi khi hủy điểm danh", "error")
    finally:
        cur.close(); conn.close()
    return redirect(url_for("attendance.detail", session_id=session_id))

# ----------------- Mobile recognition API -----------------
# Cache recognition data
_enc_loaded = False
_known_encs = []
_known_names = []
_label_map = {}
_yolo_model = None

def _load_encodings_and_labels():
    global _enc_loaded, _known_encs, _known_names, _label_map
    if _enc_loaded:
        return
    try:
        with open("encodings.pkl", "rb") as f:
            data = pickle.load(f)
            _known_encs = data.get("encodings", [])
            _known_names = data.get("names", [])
    except Exception:
        _known_encs = []; _known_names = []
    try:
        with open("labels.json", "r", encoding="utf-8") as f:
            _label_map = _json.load(f)
    except Exception:
        _label_map = {}
    _enc_loaded = True

def _get_yolo():
    global _yolo_model
    if _yolo_model is not None:
        return _yolo_model
    try:
        from ultralytics import YOLO
        _yolo_model = YOLO('yolov11n-face.pt')
        _yolo_model.overrides['verbose'] = False
    except Exception:
        _yolo_model = None
    return _yolo_model

def _recognize_frame(bgr, thresh=0.35):
    """Return (label, confidence) from a BGR image; Unknown if none.
    thresh: distance threshold (smaller is stricter). Example: 0.35 ~= 65%.
    """
    _load_encodings_and_labels()
    if not _known_encs:
        return "Unknown", 0.0
    h, w = bgr.shape[:2]
    # Giữ độ phân giải tốt hơn thay vì ép 640x480; scale về ~960px chiều rộng nếu quá lớn
    target_w = 960
    if w > target_w:
        scale = target_w / float(w)
        small = cv2.resize(bgr, (int(w*scale), int(h*scale)))
    else:
        small = bgr
    rgb_small = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
    model = _get_yolo()
    box = None
    if model is not None:
        try:
            res = model(rgb_small, verbose=False)
            for r in res:
                for bx in r.boxes:
                    conf = float(bx.conf[0])
                    if conf >= 0.5:
                        x1, y1, x2, y2 = map(int, bx.xyxy[0])
                        box = (y1, x2, y2, x1)  # top,right,bottom,left for face_recognition
                        break
        except Exception:
            box = None
    if box is None:
        try:
            locs = []
            import face_recognition
            locs = face_recognition.face_locations(rgb_small, model='hog')
            box = locs[0] if locs else None
        except Exception:
            box = None
    if box is None:
        return "Unknown", 0.0
    try:
        import face_recognition
        encs = face_recognition.face_encodings(rgb_small, [box], num_jitters=0)
        if not encs:
            return "Unknown", 0.0
        enc = encs[0]
        dists = face_recognition.face_distance(_known_encs, enc)
        idx = int(np.argmin(dists))
        min_dist = float(dists[idx])
        if min_dist <= float(thresh):
            label = _known_names[idx]
            confidence = (1 - min_dist) * 100.0
            return label, confidence
        else:
            return "Unknown", (1 - min_dist) * 100.0
    except Exception:
        return "Unknown", 0.0

def _recognize_frame_multi(bgr, thresh=0.35):
    """Return list of (label, confidence) for all faces found in the image.
    Deduplicates labels; keeps the highest confidence for each label.
    """
    _load_encodings_and_labels()
    results = []
    if not _known_encs:
        return results
    h, w = bgr.shape[:2]
    # Keep better resolution, scale down if too large
    target_w = 960
    if w > target_w:
        scale = target_w / float(w)
        small = cv2.resize(bgr, (int(w * scale), int(h * scale)))
    else:
        small = bgr
    rgb = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
    boxes = []
    # Try YOLO for multiple faces
    model = _get_yolo()
    if model is not None:
        try:
            res = model(rgb, verbose=False)
            for r in res:
                for bx in r.boxes:
                    conf = float(bx.conf[0])
                    if conf >= 0.5:
                        x1, y1, x2, y2 = map(int, bx.xyxy[0])
                        boxes.append((y1, x2, y2, x1))  # top,right,bottom,left
        except Exception:
            boxes = []
    # Fallback HOG if YOLO no faces
    if not boxes:
        try:
            import face_recognition
            boxes = face_recognition.face_locations(rgb, model='hog')
        except Exception:
            boxes = []
    if not boxes:
        return results
    try:
        import face_recognition
        encs = face_recognition.face_encodings(rgb, boxes, num_jitters=0)
        best_by_label = {}
        for enc in encs:
            dists = face_recognition.face_distance(_known_encs, enc)
            idx = int(np.argmin(dists))
            min_dist = float(dists[idx])
            if min_dist <= float(thresh):
                label = _known_names[idx]
                conf = (1 - min_dist) * 100.0
                if label not in best_by_label or conf > best_by_label[label]:
                    best_by_label[label] = conf
        results = [(lbl, c) for lbl, c in best_by_label.items()]
        # Sort by confidence desc
        results.sort(key=lambda x: x[1], reverse=True)
        return results
    except Exception:
        return []

@attendance_bp.route("/api/recognize_submit", methods=["POST"])
def api_recognize_submit():
    settings = get_all_settings()
    if settings.get("phone_attendance_enabled") != '1':
        return jsonify({"status": "error", "message": "Phone attendance disabled"}), 403
    data = request.get_json(force=True)
    session_id = data.get("session_id")
    img_data = data.get("image")
    if not session_id or not img_data:
        return jsonify({"status": "error", "message": "Thiếu tham số"}), 400
    # Validate session exists and started
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT status FROM attendance_sessions WHERE id=%s", (session_id,))
    sess = cur.fetchone()
    if not sess:
        cur.close(); conn.close();
        return jsonify({"status": "error", "message": "Phiên không tồn tại"}), 404
    if str(sess[0]) != 'started':
        cur.close(); conn.close();
        return jsonify({"status": "error", "message": "Phiên chưa bắt đầu"}), 400
    # Decode data URL
    try:
        if "," in img_data:
            img_data = img_data.split(",", 1)[1]
        img_bytes = base64.b64decode(img_data)
        arr = np.frombuffer(img_bytes, dtype=np.uint8)
        bgr = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if bgr is None:
            return jsonify({"status": "error", "message": "Ảnh không hợp lệ"}), 400
    except Exception as e:
        return jsonify({"status": "error", "message": f"Không giải mã ảnh: {e}"}), 400
    # Ngưỡng nhận diện lấy từ cấu hình: confidence >= X%  <=> distance <= 1 - X/100
    try:
        min_conf = float(settings.get("recognition_min_confidence", "65"))
    except Exception:
        min_conf = 65.0
    dist_thresh = max(0.0, min(1.0, 1.0 - (min_conf/100.0)))
    results = _recognize_frame_multi(bgr, thresh=dist_thresh)
    payload = {"status": "ok", "results": [
        {"label": lbl, "confidence": conf} for (lbl, conf) in results
    ]}
    # For backward-compatibility include top-1
    if results:
        payload["label"], payload["confidence"] = results[0][0], results[0][1]
    # Mark all recognized students
    marked = []
    try:
        for lbl, conf in results:
            cur.execute("SELECT id FROM students WHERE face_label=%s", (lbl,))
            row = cur.fetchone()
            if row:
                student_id = row[0]
                cur.execute(
                    """
                    INSERT INTO attendance_records (session_id, student_id, status, confidence, device)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (session_id, student_id) DO UPDATE SET status=EXCLUDED.status, confidence=EXCLUDED.confidence, device=EXCLUDED.device, captured_at=NOW()
                    """,
                    (session_id, student_id, 'present', conf, 'phone')
                )
                marked.append({"student_id": student_id, "label": lbl, "confidence": conf})
        conn.commit()
        payload["marked"] = marked
    except Exception as e:
        conn.rollback(); payload["error"] = str(e)
    finally:
        cur.close(); conn.close()
    return jsonify(payload)

@attendance_bp.route("/<int:session_id>/start", methods=["POST"])
@role_required("admin", "lecturer")
def start(session_id):
    conn = get_connection(); cur = conn.cursor()
    cur.execute("UPDATE attendance_sessions SET status='started', start_time=NOW() WHERE id=%s", (session_id,))
    conn.commit(); cur.close(); conn.close()
    flash("Đã bắt đầu điểm danh", "success")
    return redirect(url_for("attendance.detail", session_id=session_id))

@attendance_bp.route("/<int:session_id>/end", methods=["POST"])
@role_required("admin", "lecturer")
def end(session_id):
    conn = get_connection(); cur = conn.cursor()
    cur.execute("UPDATE attendance_sessions SET status='ended', end_time=NOW() WHERE id=%s", (session_id,))
    conn.commit(); cur.close(); conn.close()
    flash("Đã kết thúc điểm danh", "success")
    return redirect(url_for("attendance.detail", session_id=session_id))

@attendance_bp.route("/<int:session_id>/export", methods=["GET"])
@login_required
def export_session(session_id):
    """Xuất báo cáo điểm danh cho một buổi học dưới dạng Excel."""
    conn = get_connection(); cur = conn.cursor()
    # Thông tin phiên
    cur.execute(
        """
        SELECT id, class_id, subject_id, session_date, time_slot, room, status, session_code
        FROM attendance_sessions WHERE id=%s
        """,
        (session_id,)
    )
    sess = cur.fetchone()
    if not sess:
        cur.close(); conn.close()
        return jsonify({"status": "error", "message": "Không tìm thấy buổi học"}), 404
    class_id, subject_id = sess[1], sess[2]
    # Mã lớp/môn
    try:
        cur.execute("SELECT code, name FROM classes WHERE id=%s", (class_id,))
        cls_row = cur.fetchone(); class_code = cls_row[0] if cls_row else ""
    except Exception:
        class_code = ""
    try:
        cur.execute("SELECT code, name FROM subjects WHERE id=%s", (subject_id,))
        sub_row = cur.fetchone(); subject_code = sub_row[0] if sub_row else ""
    except Exception:
        subject_code = ""
    # Danh sách sinh viên theo lớp/môn
    has_subject_students = False
    try:
        cur.execute("SELECT EXISTS(SELECT 1 FROM student_subjects WHERE subject_id=%s)", (subject_id,))
        row = cur.fetchone(); has_subject_students = bool(row[0]) if row else False
    except Exception:
        has_subject_students = False
    if has_subject_students:
        try:
            cur.execute(
                """
                SELECT s.id, s.student_code, s.fullname
                FROM student_subjects ss
                JOIN students s ON s.id=ss.student_id
                WHERE ss.subject_id=%s AND s.class_id=%s
                ORDER BY s.fullname
                """,
                (subject_id, class_id)
            )
            students = cur.fetchall()
        except Exception:
            cur.execute("SELECT s.id, s.student_code, s.fullname FROM students s WHERE s.class_id=%s ORDER BY s.fullname", (class_id,))
            students = cur.fetchall()
    else:
        cur.execute("SELECT s.id, s.student_code, s.fullname FROM students s WHERE s.class_id=%s ORDER BY s.fullname", (class_id,))
        students = cur.fetchall()
    # Bản ghi điểm danh
    cur.execute("SELECT student_id, status, confidence, device, captured_at FROM attendance_records WHERE session_id=%s", (session_id,))
    recs = cur.fetchall(); cur.close(); conn.close()
    rec_map = {r[0]: r for r in recs}

    # Tạo workbook
    wb = Workbook()
    sheet_title = f"{class_code}-{subject_code}" if (class_code or subject_code) else "Diem danh"
    # Tạo sheet mới để đảm bảo kiểu Worksheet, tránh Optional/Chartsheet
    ws: Worksheet = wb.create_sheet(title=sheet_title)
    # Xóa sheet mặc định nếu còn tồn tại
    default_ws = wb.active
    if default_ws is not None and default_ws is not ws:
        try:
            wb.remove(default_ws)
        except Exception:
            pass

    # Thông tin phiên ở đầu file
    headers_info = [
        ("Lớp", class_code),
        ("Môn", subject_code),
        ("Ngày", str(sess[3])),
        ("Ca", str(sess[4] or "")),
        ("Phòng", str(sess[5] or "")),
        ("Trạng thái", str(sess[6])),
        ("Mã phiên", str(sess[7])),
    ]
    row_idx = 1
    for label, value in headers_info:
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    # Dòng trống
    row_idx += 1

    # Header bảng dữ liệu
    cols = ["STT", "Mã SV", "Họ tên", "Trạng thái", "Độ tin cậy", "Thiết bị", "Thời gian"]
    for c, h in enumerate(cols, 1):
        ws.cell(row=row_idx, column=c, value=h)
    row_idx += 1

    # Ghi dữ liệu từng sinh viên
    for i, s in enumerate(students, 1):
        sid, scode, sname = s[0], s[1], s[2]
        r = rec_map.get(sid)
        status = r[1] if r else "chưa điểm danh"
        confidence = r[2] if r else None
        device = r[3] if r else ""
        captured = str(r[4]) if r else ""
        ws.cell(row=row_idx, column=1, value=i)
        ws.cell(row=row_idx, column=2, value=scode)
        ws.cell(row=row_idx, column=3, value=sname)
        ws.cell(row=row_idx, column=4, value=status)
        ws.cell(row=row_idx, column=5, value=confidence if confidence is not None else "")
        ws.cell(row=row_idx, column=6, value=device)
        ws.cell(row=row_idx, column=7, value=captured)
        row_idx += 1

    # Điều chỉnh độ rộng cột đơn giản (tránh MergedCell, dùng get_column_letter)
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            try:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Trả về file Excel
    import io
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"diem_danh_{sess[7] or session_id}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

@attendance_bp.route("/<int:session_id>/mark", methods=["POST"])
@role_required("admin", "lecturer")
def mark(session_id):
    student_id = request.form.get("student_id")
    status = request.form.get("status")
    confidence = request.form.get("confidence") or None
    device = request.form.get("device") or "manual"
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO attendance_records (session_id, student_id, status, confidence, device)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (session_id, student_id) DO UPDATE SET status=EXCLUDED.status, confidence=EXCLUDED.confidence, device=EXCLUDED.device, captured_at=NOW()
        """, (session_id, student_id, status, confidence, device))
        conn.commit(); flash("Đã cập nhật điểm danh", "success")
    except Exception:
        conn.rollback(); flash("Lỗi cập nhật điểm danh", "error")
    finally:
        cur.close(); conn.close()
    return redirect(url_for("attendance.detail", session_id=session_id))

# Điểm danh theo nhãn khuôn mặt (face_label) dành cho webcam
@attendance_bp.route("/<int:session_id>/mark_by_label", methods=["POST"])
@role_required("admin", "lecturer")
def mark_by_label(session_id):
    # Hỗ trợ cả JSON và form-urlencoded
    data = None
    try:
        data = request.get_json(silent=True, force=False)
    except Exception:
        data = None
    label = None
    status = "present"
    confidence = None
    device = "webcam"
    if data:
        label = (data.get("label") or "").strip()
        status = (data.get("status") or status).strip()
        confidence = data.get("confidence")
    else:
        label = (request.form.get("label") or "").strip()
        status = (request.form.get("status") or status).strip()
        confidence = request.form.get("confidence") or None
    if not label:
        return jsonify({"status": "error", "message": "Thiếu label"}), 400
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("SELECT id FROM students WHERE face_label=%s", (label,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify({"status": "error", "message": "Không tìm thấy sinh viên cho label này"}), 404
        student_id = row[0]
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO attendance_records (session_id, student_id, status, confidence, device)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (session_id, student_id) DO UPDATE SET status=EXCLUDED.status, confidence=EXCLUDED.confidence, device=EXCLUDED.device, captured_at=NOW()
            """,
            (session_id, student_id, status, confidence, device)
        )
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status": "ok"})
    except Exception as e:
        conn.rollback()
        try:
            cur.close()
            conn.close()
        except Exception:
            pass
        return jsonify({"status": "error", "message": str(e)}), 500

# Mobile join page (simple)
@attendance_bp.route("/mobile/<session_code>")
def mobile_join(session_code):
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT id, class_id, subject_id, session_date, status FROM attendance_sessions WHERE session_code=%s", (session_code,))
    sess = cur.fetchone()
    if not sess:
        cur.close(); conn.close()
        return "Phiên không tồn tại", 404
    # Lấy danh sách sinh viên theo lớp/môn
    class_id = sess[1]; subject_id = sess[2]
    has_subject_students = False
    try:
        cur.execute("SELECT EXISTS(SELECT 1 FROM student_subjects WHERE subject_id=%s)", (subject_id,))
        row = cur.fetchone(); has_subject_students = bool(row[0]) if row else False
    except Exception:
        has_subject_students = False
    if has_subject_students:
        try:
            cur.execute(
                """
                SELECT s.id, s.student_code, s.fullname
                FROM student_subjects ss
                JOIN students s ON s.id=ss.student_id
                WHERE ss.subject_id=%s AND s.class_id=%s
                ORDER BY s.fullname
                """,
                (subject_id, class_id)
            )
            students = cur.fetchall()
        except Exception:
            cur.execute("SELECT s.id, s.student_code, s.fullname FROM students s WHERE s.class_id=%s ORDER BY s.fullname", (class_id,))
            students = cur.fetchall()
    else:
        cur.execute("SELECT s.id, s.student_code, s.fullname FROM students s WHERE s.class_id=%s ORDER BY s.fullname", (class_id,))
        students = cur.fetchall()
    # Đọc cấu hình để biết có bật điểm danh bằng điện thoại không
    settings = get_all_settings()
    phone_enabled = (settings.get("phone_attendance_enabled") == '1')
    cur.close(); conn.close()
    return render_template("attendance_mobile.html", session=sess, students=students, phone_enabled=phone_enabled)

# Trang nhập mã phiên (dùng khi không quét QR)
@attendance_bp.route("/mobile", methods=["GET", "POST"])
def mobile_enter():
    error = None
    if request.method == "POST":
        code = (request.form.get("session_code") or "").strip()
        if not code:
            error = "Vui lòng nhập mã phiên"
        else:
            conn = get_connection(); cur = conn.cursor()
            try:
                cur.execute("SELECT session_code FROM attendance_sessions WHERE session_code=%s", (code,))
                row = cur.fetchone()
                if row:
                    return redirect(url_for("attendance.mobile_join", session_code=code))
                else:
                    error = "Mã phiên không tồn tại"
            finally:
                cur.close(); conn.close()
    return render_template("attendance_mobile_enter.html", error=error)

# API to submit attendance from mobile (future camera integration)
@attendance_bp.route("/api/submit", methods=["POST"])
def api_submit():
    data = request.get_json(force=True)
    session_id = data.get("session_id")
    student_id = data.get("student_id")
    status = data.get("status", "present")
    confidence = data.get("confidence")
    device = "phone"
    conn = get_connection(); cur = conn.cursor()
    try:
        # Validate session exists and started
        cur.execute("SELECT status FROM attendance_sessions WHERE id=%s", (session_id,))
        sess = cur.fetchone()
        if not sess:
            cur.close(); conn.close();
            return jsonify({"status": "error", "message": "Phiên không tồn tại"}), 404
        if str(sess[0]) != 'started':
            cur.close(); conn.close();
            return jsonify({"status": "error", "message": "Phiên chưa bắt đầu"}), 400
        cur.execute("""
            INSERT INTO attendance_records (session_id, student_id, status, confidence, device)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (session_id, student_id) DO UPDATE SET status=EXCLUDED.status, confidence=EXCLUDED.confidence, device=EXCLUDED.device, captured_at=NOW()
        """, (session_id, student_id, status, confidence, device))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status": "ok"})
    except Exception as e:
        conn.rollback(); cur.close(); conn.close()
        return jsonify({"status": "error", "message": str(e)}), 500
