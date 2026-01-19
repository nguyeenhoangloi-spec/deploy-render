from flask import Blueprint, render_template, request, redirect, url_for, flash
from connect_postgres import get_connection
from auth import login_required, role_required

students_bp = Blueprint("students", __name__)

@students_bp.route("/")
@login_required
def index():
    q = (request.args.get("q") or "").strip()
    class_id_raw = request.args.get("class_id") or ""
    class_id = int(class_id_raw) if class_id_raw.isdigit() else None
    conn = get_connection()
    cur = conn.cursor()
    try:
        # Load classes for filter
        cur.execute("SELECT id, code, name FROM classes ORDER BY code")
        classes = cur.fetchall()
        # Build query
        sql = "SELECT s.id, s.student_code, s.fullname, c.code, c.name, s.email, s.phone, s.face_label FROM students s LEFT JOIN classes c ON s.class_id=c.id"
        params = []
        where = []
        if q:
            where.append("(LOWER(s.fullname) LIKE %s OR LOWER(s.student_code) LIKE %s)")
            params.extend([f"%{q.lower()}%", f"%{q.lower()}%"])
        if class_id is not None:
            where.append("s.class_id = %s")
            params.append(class_id)
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY s.fullname"
        cur.execute(sql, params)
        raw_rows = cur.fetchall()
        # Shape rows into dicts for stable template fields
        rows = [
            {
                "id": r[0],
                "student_code": r[1],
                "fullname": r[2],
                "class_code": r[3] or "",
                "class_name": r[4] or "",
                "email": r[5] or "",
                "phone": r[6] or "",
                "face_label": r[7] or "",
            }
            for r in raw_rows
        ]
    except Exception:
        # Fallback an toàn nếu có lỗi truy vấn
        rows = []
        classes = []
    finally:
        cur.close(); conn.close()
    # Truyền lại class_id_raw để giữ giá trị chọn trong select
    return render_template("students_list.html", rows=rows, classes=classes, q=q, class_id=class_id_raw)


@students_bp.route("/api")
@login_required
def api_index():
    """Diagnostic: return students and classes as JSON to verify DB and route."""
    q = (request.args.get("q") or "").strip()
    class_id_raw = request.args.get("class_id") or ""
    class_id = int(class_id_raw) if class_id_raw.isdigit() else None
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("SELECT id, code, name FROM classes ORDER BY code")
        classes = cur.fetchall()
        sql = "SELECT s.id, s.student_code, s.fullname, c.code, c.name, s.email, s.phone, s.face_label FROM students s LEFT JOIN classes c ON s.class_id=c.id"
        params = []; where = []
        if q:
            where.append("(LOWER(s.fullname) LIKE %s OR LOWER(s.student_code) LIKE %s)")
            params.extend([f"%{q.lower()}%", f"%{q.lower()}%"])
        if class_id is not None:
            where.append("s.class_id = %s"); params.append(class_id)
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY s.fullname"
        cur.execute(sql, params); rows = cur.fetchall()
        return {"status": "ok", "rows": rows, "classes": classes}
    except Exception as e:
        return {"status": "error", "message": str(e)}, 500
    finally:
        cur.close(); conn.close()

@students_bp.route("/new", methods=["GET", "POST"])
@role_required("admin", "lecturer")
def create():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, code, name FROM classes ORDER BY code")
    classes = cur.fetchall()
    if request.method == "POST":
        student_code = request.form.get("student_code")
        fullname = request.form.get("fullname")
        email = request.form.get("email") or None
        phone = request.form.get("phone") or None
        class_id = request.form.get("class_id") or None
        try:
            cur.execute(
                "INSERT INTO students (student_code, fullname, email, phone, class_id) VALUES (%s, %s, %s, %s, %s)",
                (student_code, fullname, email, phone, class_id)
            )
            conn.commit()
            flash("Đã thêm sinh viên", "success")
            return redirect(url_for("students.index"))
        except Exception as e:
            conn.rollback()
            msg = str(e)
            if "unique" in msg.lower() and "student_code" in msg.lower():
                flash("Mã sinh viên đã tồn tại", "error")
            else:
                flash(f"Lỗi thêm sinh viên: {e}", "error")
            # Trả lại form với dữ liệu đã nhập
            form = {
                "student_code": student_code,
                "fullname": fullname,
                "email": email,
                "phone": phone,
                "class_id": (int(class_id) if class_id and class_id.isdigit() else None),
            }
            cur.close(); conn.close()
            return render_template("students_form.html", classes=classes, form=form)
    cur.close()
    conn.close()
    return render_template("students_form.html", classes=classes, form={})

@students_bp.route("/<int:student_id>/edit", methods=["GET", "POST"])
@role_required("admin", "lecturer")
def edit(student_id):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, code, name FROM classes ORDER BY code")
    classes = cur.fetchall()
    cur.execute("SELECT id, student_code, fullname, email, phone, class_id, face_label FROM students WHERE id=%s", (student_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return redirect(url_for("students.index"))
    if request.method == "POST":
        student_code = request.form.get("student_code")
        fullname = request.form.get("fullname")
        email = request.form.get("email") or None
        phone = request.form.get("phone") or None
        class_id = request.form.get("class_id") or None
        try:
            cur.execute(
                "UPDATE students SET student_code=%s, fullname=%s, email=%s, phone=%s, class_id=%s WHERE id=%s",
                (student_code, fullname, email, phone, class_id, student_id)
            )
            conn.commit()
            flash("Đã cập nhật", "success")
            return redirect(url_for("students.index"))
        except Exception as e:
            conn.rollback()
            msg = str(e)
            if "unique" in msg.lower() and "student_code" in msg.lower():
                flash("Mã sinh viên đã tồn tại", "error")
            else:
                flash(f"Lỗi cập nhật: {e}", "error")
            # Hiển thị lại form với giá trị đã nhập
            form = {
                "id": student_id,
                "student_code": student_code,
                "fullname": fullname,
                "email": email,
                "phone": phone,
                "class_id": (int(class_id) if class_id and class_id.isdigit() else None),
                "face_label": None,
            }
            cur.close(); conn.close()
            return render_template("students_form.html", classes=classes, form=form)
    cur.close(); conn.close()
    form = {"id": row[0], "student_code": row[1], "fullname": row[2], "email": row[3], "phone": row[4], "class_id": row[5], "face_label": row[6]}
    return render_template("students_form.html", classes=classes, form=form)

@students_bp.route("/<int:student_id>/delete", methods=["POST"])
@role_required("admin")
def delete(student_id):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM students WHERE id=%s", (student_id,))
        conn.commit()
        flash("Đã xóa", "success")
    except Exception:
        conn.rollback()
        flash("Lỗi xóa", "error")
    finally:
        cur.close(); conn.close()
    return redirect(url_for("students.index"))


@students_bp.route("/<int:student_id>/capture_face")
@role_required("admin", "lecturer")
def capture_face(student_id):
    """Mở trang chụp ảnh khuôn mặt đã điền sẵn họ tên và lớp."""
    conn = get_connection(); cur = conn.cursor()
    cur.execute(
        """
        SELECT s.fullname, COALESCE(c.code, '') AS class_code
        FROM students s
        LEFT JOIN classes c ON s.class_id = c.id
        WHERE s.id=%s
        """,
        (student_id,)
    )
    row = cur.fetchone(); cur.close(); conn.close()
    if not row:
        flash("Không tìm thấy sinh viên", "error")
        return redirect(url_for("students.index"))
    fullname, class_code = row[0], row[1]
    # Chuyển hướng sang trang chụp với query để tự điền form
    return redirect(url_for("add_user.capture_page", fullname=fullname, classname=class_code))

# Excel import
@students_bp.route("/import", methods=["GET", "POST"])
@role_required("admin", "lecturer")
def import_excel():
    from openpyxl import load_workbook
    from io import BytesIO
    msg = ""
    if request.method == "POST" and "file" in request.files:
        f = request.files["file"]
        try:
            # Đọc dữ liệu Excel từ nội dung file upload an toàn cho Pylance
            wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
            ws = wb.active if hasattr(wb, "active") else None
            if ws is None and hasattr(wb, "worksheets") and wb.worksheets:
                ws = wb.worksheets[0]
            conn = get_connection(); cur = conn.cursor()
            inserted = 0
            if ws is None:
                raise ValueError("File Excel không có sheet hợp lệ")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # header
                # Bỏ qua dòng thiếu dữ liệu
                if not row or len(row) < 5:
                    continue
                # Ép kiểu và loại bỏ khoảng trắng dư thừa
                student_code = str(row[0]).strip() if row[0] is not None else None
                fullname = str(row[1]).strip() if row[1] is not None else None
                email = str(row[2]).strip() if row[2] is not None else None
                phone = str(row[3]).strip() if row[3] is not None else None
                class_code = str(row[4]).strip() if row[4] is not None else None
                # Nếu thiếu mã SV hoặc họ tên thì bỏ qua
                if not student_code or not fullname:
                    continue
                class_id = None
                code_val, name_val = None, None
                if class_code:
                    # Tách theo dấu '-' nếu có, loại bỏ khoảng trắng
                    parts = [p.strip() for p in class_code.split('-')]
                    if len(parts) == 2:
                        code_val, name_val = parts[0], parts[1]
                    elif len(parts) == 1:
                        code_val = parts[0]
                    # Tìm class_id theo code và name nếu có
                    if code_val and name_val:
                        cur.execute("SELECT id FROM classes WHERE code=%s AND name=%s", (code_val, name_val))
                        found = cur.fetchone()
                        if found:
                            class_id = found[0]
                    if class_id is None and code_val:
                        cur.execute("SELECT id FROM classes WHERE code=%s", (code_val,))
                        found = cur.fetchone()
                        if found:
                            class_id = found[0]
                cur.execute(
                    "INSERT INTO students (student_code, fullname, email, phone, class_id) VALUES (%s, %s, %s, %s, %s) "
                    "ON CONFLICT (student_code) DO UPDATE SET fullname=EXCLUDED.fullname, email=EXCLUDED.email, phone=EXCLUDED.phone, class_id=EXCLUDED.class_id",
                    (student_code, fullname, email if email else None, phone if phone else None, class_id)
                )
                inserted += 1
            conn.commit(); cur.close(); conn.close()
            msg = f"Đã nhập {inserted} dòng"
        except Exception as e:
            msg = f"Lỗi nhập: {e}"
    return render_template("students_import.html", message=msg)
